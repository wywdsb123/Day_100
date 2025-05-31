import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 数据准备
student_names = ["关羽", "张飞", "赵云", "马超"]
scores = [[random.randrange(50, 101) for _ in range(3)] for _ in range(4)]
titles = ("姓名", "语文", "数学", "英语")

# 创建工作簿
wb = Workbook()
sheet = wb.active
sheet.title = "一年级二班"

# 设置标题样式
header_font = Font(name="华文楷体", bold=True, size=12, color="FF0000")
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 写入标题
for col, title in enumerate(titles, 1):
    cell = sheet.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 写入数据
for row_idx, (name, score_row) in enumerate(zip(student_names, scores), 2):
    sheet.cell(row=row_idx, column=1, value=name)
    for col_idx, score in enumerate(score_row, 2):
        sheet.cell(row=row_idx, column=col_idx, value=score)

# 保存文件
wb.save("考试成绩表.xlsx")