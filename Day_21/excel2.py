import random
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "期末成绩"

titles =("姓名","语文","数学","英语")
for col_index,title in enumerate(titles):
    sheet.cell(1,col_index+1,title)

names = ("关羽","张飞","赵云","马超","黄忠")
for row_index,name in enumerate(names):
    sheet.cell(row_index+2,1,name)
    for col_index in range(2,5):
        sheet.cell(row_index+2,col_index,random.randrange(50,100))
wb.save("考试成绩表.xlsx")