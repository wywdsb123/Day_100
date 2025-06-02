"""wang
"""
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side

from excel import sheet

alignment =Alignment(horizontal="center",vertical="center")
side =Side(color="ff7f50",style="mediumDashed")

wb = openpyxl.load_workbook("考试成绩表.xlsx")
sheet = wb.worksheets[0]

sheet.row_dimensions[1].height=30
sheet.column_dimensions["E"].width = 30

sheet["E1"]="平均分"
sheet.cell(1,5).font = Font(size=18,color="ff1493",name="华文楷体")
sheet.cell(1,5).alignment =alignment

sheet.cell(1,5).border=Border(left=side,top=side,right=side,bottom=side)
for i in range(2,7):
    sheet[f"E{i}"]=f"=average(B{i}:D{i})"
    sheet.cell(i,5).font =Font(size=12,color="4169e1",italic="True")
    sheet.cell(i,5).alignment = alignment
wb.save("考试成绩表.xlsx")