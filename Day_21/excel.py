import datetime
import openpyxl
wb= openpyxl.load_workbook("C:\\Users\\wywdsb\\Desktop\\阿里巴巴2020年股票数据.xlsx")
print(wb.sheetnames)
sheet = wb.worksheets[0]
print(sheet.dimensions)
print(sheet.max_row,sheet.max_column)

print(sheet.cell(3,3,).value)
print(sheet["C3"].value)
print(sheet["G255"])

print(sheet["A2:C5"])
for row_ch in range(2,sheet.max_row+1):
    for col_ch in  "ABCDEFG":
        value = sheet[f"{col_ch}{row_ch}"].value
        if type(value) == datetime.datetime:
            print(value.strftime("%Y年%m月%d日"),end='\t')
        elif type(value) == int:
            print(f"{value:<10d}",end="\t")
        elif type(value) == float:
            print(f"{value:.4f}",end = "\t")
        else:
            print(value,end="\t")
    print()
